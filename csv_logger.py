import os
from datetime import datetime, date

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from config_manager import DATA_DIR

import time


def _safe_save(wb, filepath, retries=3, delay=2):
    """安全保存，遇到文件锁定时自动重试"""
    for i in range(retries):
        try:
            wb.save(filepath)
            return True
        except PermissionError:
            if i < retries - 1:
                time.sleep(delay)
            else:
                print(f"[警告] 无法保存 {os.path.basename(filepath)}，文件可能被 Excel 占用")
                return False
        except Exception as e:
            print(f"[警告] 保存失败: {e}")
            return False
    return False

WRITE_LOG_XLSX = os.path.join(DATA_DIR, "write_log.xlsx")
DAILY_XLSX = os.path.join(DATA_DIR, "daily_summary.xlsx")
TREND_XLSX = os.path.join(DATA_DIR, "trend_summary.xlsx")

WRITE_LOG_HEADER = [
    "时间戳(timestamp)", "磁盘(disk)", "累计写入GB(total_written_gb)", "累计写入TB(total_written_tb)",
    "本次增量GB(delta_gb)", "今日增量GB(today_gb)", "温度(temperature_c)",
    "写入进程1(top1_process)", "写入进程2(top2_process)", "写入进程3(top3_process)"
]

DAILY_HEADER = [
    "日期(date)", "磁盘(disk)", "写入量GB(written_gb)", "写入量TB(written_tb)",
    "日初累计TB(day_start_tb)", "日末累计TB(day_end_tb)"
]

TREND_HEADER = [
    "日期(date)", "写入量GB(written_gb)", "写入量TB(written_tb)", "累计写入TB(cumulative_tb)",
    "近7天日均GB(avg_7d_gb)", "采集次数(readings)"
]


def _auto_width(ws):
    """自动调整所有列宽"""
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                val = str(cell.value) if cell.value else ""
                # 中文字符算 2 个宽度
                width = sum(2 if ord(c) > 127 else 1 for c in val)
                max_len = max(max_len, width)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def _ensure_xlsx(filepath, header):
    """确保 xlsx 存在且有表头，返回 workbook"""
    if os.path.exists(filepath):
        wb = load_workbook(filepath)
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(header)
        ws.freeze_panes = "A2"
        wb.save(filepath)
    return wb


def _read_all_rows(filepath):
    """读取所有数据行（不含表头）"""
    if not os.path.exists(filepath):
        return []
    wb = load_workbook(filepath)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(cell is not None for cell in row):
            rows.append(list(row))
    return rows


# ─── 对外接口 ───
def _format_source(s):
    """格式化单个进程来源: 'chrome.exe(120.5MB)'"""
    mb = s["delta_bytes"] / (1024 ** 2)
    if mb >= 1024:
        return f'{s["name"]}({mb/1024:.1f}GB)'
    return f'{s["name"]}({mb:.1f}MB)'


def log_write(timestamp, disk, total_bytes, delta_bytes, today_bytes, temp, sources=None):
    """写入详细记录

    Args:
        sources: 可选，[{"name": str, "pid": int, "delta_bytes": int}, ...] 最多取前3个
    """
    wb = _ensure_xlsx(WRITE_LOG_XLSX, WRITE_LOG_HEADER)
    ws = wb.active

    total_gb = total_bytes / (1024 ** 3)
    total_tb = total_bytes / (1024 ** 4)
    delta_gb = delta_bytes / (1024 ** 3)
    today_gb = today_bytes / (1024 ** 3)

    # 进程来源列（最多3列，从左到右递减）
    source_cols = ["", "", ""]
    if sources:
        for i, s in enumerate(sources[:3]):
            source_cols[i] = _format_source(s)

    ws.append([
        timestamp.strftime("%Y-%m-%d %H:%M:%S"),
        disk,
        round(total_gb, 2),
        round(total_tb, 3),
        round(delta_gb, 2),
        round(today_gb, 2),
        temp if temp is not None else "",
        source_cols[0],
        source_cols[1],
        source_cols[2],
    ])

    _auto_width(ws)
    _safe_save(wb, WRITE_LOG_XLSX)


def log_daily_summary(disk, written_bytes, day_start_bytes, day_end_bytes):
    """写入每日汇总"""
    wb = _ensure_xlsx(DAILY_XLSX, DAILY_HEADER)
    ws = wb.active

    today = date.today().isoformat()
    written_gb = written_bytes / (1024 ** 3)
    written_tb = written_bytes / (1024 ** 4)
    start_tb = day_start_bytes / (1024 ** 4)
    end_tb = day_end_bytes / (1024 ** 4)

    ws.append([
        today, disk,
        round(written_gb, 2), round(written_tb, 3),
        round(start_tb, 3), round(end_tb, 3)
    ])

    _auto_width(ws)
    _safe_save(wb, DAILY_XLSX)


def log_trend(written_bytes, cumulative_bytes, readings):
    """更新趋势统计（同一天更新，跨天追加）"""
    wb = _ensure_xlsx(TREND_XLSX, TREND_HEADER)
    ws = wb.active

    today = date.today().isoformat()
    written_gb = written_bytes / (1024 ** 3)
    written_tb = written_bytes / (1024 ** 4)
    cum_tb = cumulative_bytes / (1024 ** 4)

    # 查找今天是否已有记录
    today_row = None
    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row_idx, 1).value == today:
            today_row = row_idx
            break

    # 计算近7天日均
    recent_gb = [written_gb]
    for row_idx in range(max(2, ws.max_row - 5), ws.max_row + 1):
        if today_row and row_idx == today_row:
            continue
        val = ws.cell(row_idx, 2).value
        if val is not None:
            try:
                recent_gb.append(float(val))
            except (ValueError, TypeError):
                pass
    avg_7d = sum(recent_gb) / len(recent_gb) if recent_gb else 0

    if today_row:
        # 更新今天的行
        ws.cell(today_row, 2, round(written_gb, 2))
        ws.cell(today_row, 3, round(written_tb, 3))
        ws.cell(today_row, 4, round(cum_tb, 3))
        ws.cell(today_row, 5, round(avg_7d, 2))
        ws.cell(today_row, 6, readings)
    else:
        ws.append([
            today,
            round(written_gb, 2), round(written_tb, 3),
            round(cum_tb, 3), round(avg_7d, 2), readings
        ])

    _auto_width(ws)
    _safe_save(wb, TREND_XLSX)


def get_today_stats():
    """获取今日统计数据"""
    if not os.path.exists(TREND_XLSX):
        return None
    today = date.today().isoformat()
    wb = load_workbook(TREND_XLSX)
    ws = wb.active
    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row_idx, 1).value == today:
            return {
                "date": ws.cell(row_idx, 1).value,
                "written_gb": float(ws.cell(row_idx, 2).value or 0),
                "written_tb": float(ws.cell(row_idx, 3).value or 0),
                "cumulative_tb": float(ws.cell(row_idx, 4).value or 0),
                "avg_7d_gb": float(ws.cell(row_idx, 5).value or 0),
                "readings": int(ws.cell(row_idx, 6).value or 0),
            }
    return None


def get_recent_trend(days=7):
    """获取最近 N 天趋势"""
    if not os.path.exists(TREND_XLSX):
        return []
    wb = load_workbook(TREND_XLSX)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(cell is not None for cell in row):
            rows.append(list(row))
    return rows[-days:]
